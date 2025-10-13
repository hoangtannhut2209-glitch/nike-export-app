# -*- coding: utf-8 -*-
"""
Template routes for Nike Export Web App - FIXED VERSION
"""

from flask import Blueprint, render_template, request, jsonify, send_file, current_app, flash, redirect, url_for
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime
from openpyxl import load_workbook

from core.template_manager_nike import NikeTemplateManager, create_sample_nike_template
from core.pdf_processor import get_invoice_data_dict
from database import get_invoice_by_number, get_extracted_data

templates_bp = Blueprint('templates', __name__)

@templates_bp.route('/')
def list_templates():
    """List all available templates"""
    template_manager = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER'])
    template_dir = current_app.config['TEMPLATE_FOLDER']
    os.makedirs(template_dir, exist_ok=True)

    templates = []
    # Only show sheets from MergeDataNike.xlsm (except Data)
    merge_file = os.path.join(template_dir, 'MergeDataNike.xlsm')
    if os.path.exists(merge_file):
        info = template_manager.get_template_info(merge_file)
        for sheet_name in info['sheets']:
            if sheet_name.strip() != 'Data':
                sheet_info = info.copy()
                sheet_info['name'] = sheet_name
                sheet_info['display_name'] = sheet_name.strip()
                sheet_info['sheet_name'] = sheet_name
                sheet_info['type'] = 'Nike Form'
                templates.append(sheet_info)

    # If no MergeDataNike.xlsm or no sheets, show nothing
    # (Optionally, could show a message in template)
    
    return render_template('templates/list.html', templates=templates)

@templates_bp.route('/<template_name>')
def template_detail(template_name):
    """Template detail page"""
    template_manager = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER'])
    
    # Always use MergeDataNike.xlsm and find the real sheet name (case-insensitive, strip spaces)
    merge_file = os.path.join(current_app.config['TEMPLATE_FOLDER'], 'MergeDataNike.xlsm')
    sheet_name = None
    if os.path.exists(merge_file):
        info = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER']).get_template_info(merge_file)
        for s in info['sheets']:
            if s.strip().lower() == template_name.strip().lower():
                sheet_name = s
                break
    if sheet_name:
        template_path = merge_file
    else:
        return redirect(url_for('templates.list_templates'))
    
    if not os.path.exists(template_path):
        flash('Template not found', 'error')
        return redirect(url_for('templates.list_templates'))
    
    # Validate template
    errors, warnings = template_manager.validate_template(template_path)
    
    if errors:
        flash(f'Template validation failed: {"; ".join(errors)}', 'error')
        return redirect(url_for('templates.list_templates'))
    
    # Get template info
    template_info = template_manager.get_template_info(template_path)
    
    # Add sheet-specific info
    template_info['display_name'] = template_name
    template_info['sheet_name'] = sheet_name
    template_info['is_nike_form'] = sheet_name is not None
    
    return render_template('templates/detail.html', 
                         template=template_info, 
                         errors=errors, 
                         warnings=warnings,
                         chr=chr)

@templates_bp.route('/template/download/<template_name>')
def download_original_template(template_name):
    """Download template file"""
    # Check if this is a Nike Form sheet
    if template_name in ['CPTPP CANADA', 'CPTPP MEXICO', 'D', 'AANZ', 'AK', 'CONG VAN ASI', 'CONG VAN CTH']:
        template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], 'MergeDataNike.xlsm')
        download_name = f"{template_name}.xlsm"
    else:
        template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], template_name)
        download_name = template_name
    
    if not os.path.exists(template_path):
        flash('Template not found', 'error')
        return redirect(url_for('templates.list_templates'))
    
    return send_file(template_path, as_attachment=True, download_name=download_name)

@templates_bp.route('/preview/<template_name>')
def preview_template(template_name):
    """Preview template content with original formatting - FIXED VERSION"""
    try:
        current_app.logger.info(f"Preview request for template: '{template_name}'")
        
        merge_file = os.path.join(current_app.config['TEMPLATE_FOLDER'], 'MergeDataNike.xlsm')
        current_app.logger.info(f"Looking for file: {merge_file}")
        
        if not os.path.exists(merge_file):
            current_app.logger.error(f"File not found: {merge_file}")
            return jsonify({'success': False, 'message': 'MergeDataNike.xlsm not found'}), 404
        
        # Find actual sheet name (handle spaces and case differences)
        sheet_name = None
        try:
            wb = load_workbook(merge_file, data_only=False)
            available_sheets = wb.sheetnames
            current_app.logger.info(f"Available sheets: {available_sheets}")
            
            for s in available_sheets:
                if s.strip().lower() == template_name.strip().lower():
                    sheet_name = s
                    current_app.logger.info(f"Found matching sheet: '{s}'")
                    break
            wb.close()
        except Exception as e:
            current_app.logger.error(f"Error reading sheets: {e}")
            return jsonify({'success': False, 'message': f'Error reading file: {str(e)}'}), 500
        
        if not sheet_name:
            current_app.logger.error(f"Sheet '{template_name}' not found in available sheets: {available_sheets}")
            return jsonify({'success': False, 'message': f'Sheet not found: {template_name}'}), 404
        
        # Now load the workbook and process the sheet
        current_app.logger.info(f"Loading workbook: {merge_file}")
        wb = load_workbook(merge_file, data_only=False)
        
        # Select the worksheet
        ws = wb[sheet_name]
        current_app.logger.info(f"Processing sheet: '{sheet_name}', dimensions: {ws.max_row}x{ws.max_column}")
        
        # Get actual used range (limit to reasonable size for preview)
        max_row = min(ws.max_row, 50)  # Limit to 50 rows for preview
        max_col = min(ws.max_column, 20)  # Limit to 20 columns for preview
        
        # Extract preview data with formatting
        preview_data = []
        
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value if cell.value is not None else ""
                
                # Get safe cell formatting
                try:
                    has_border = False
                    if cell.border:
                        has_border = bool(
                            (cell.border.left and cell.border.left.style) or 
                            (cell.border.right and cell.border.right.style) or 
                            (cell.border.top and cell.border.top.style) or 
                            (cell.border.bottom and cell.border.bottom.style)
                        )
                    
                    alignment_val = 'general'
                    if cell.alignment and cell.alignment.horizontal:
                        alignment_val = str(cell.alignment.horizontal)
                    
                    cell_style = {
                        'font_bold': bool(cell.font.bold if cell.font else False),
                        'font_size': int(cell.font.size if cell.font and cell.font.size else 11),
                        'border': has_border,
                        'alignment': alignment_val,
                        'merged': False
                    }
                    
                    # Check if cell is merged (safely)
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            cell_style['merged'] = True
                            break
                            
                except Exception as style_e:
                    current_app.logger.warning(f"Style error for cell {cell.coordinate}: {style_e}")
                    cell_style = {'font_bold': False, 'font_size': 11, 'border': False, 'alignment': 'general', 'merged': False}
                
                # Check if it's a placeholder
                is_placeholder = isinstance(cell_value, str) and '{' in cell_value and '}' in cell_value
                
                row_data.append({
                    'value': str(cell_value),
                    'is_placeholder': is_placeholder,
                    'address': cell.coordinate,
                    'style': cell_style,
                    'row': row_idx,
                    'col': col_idx
                })
            preview_data.append(row_data)
        
        # Get print area and page setup info (safely)
        print_area = None
        page_margins = None
        try:
            print_area = ws.print_area if hasattr(ws, 'print_area') else None
            if hasattr(ws, 'page_margins') and ws.page_margins:
                page_margins = {
                    'left': ws.page_margins.left,
                    'right': ws.page_margins.right, 
                    'top': ws.page_margins.top,
                    'bottom': ws.page_margins.bottom
                }
        except Exception as margin_e:
            current_app.logger.warning(f"Error getting page info: {margin_e}")
        
        wb.close()
        
        current_app.logger.info(f"Preview successful for '{sheet_name}': {len(preview_data)} rows")
        
        return jsonify({
            'success': True,
            'template_name': template_name,
            'sheet_name': sheet_name,
            'preview_data': preview_data,
            'total_rows': max_row,
            'total_cols': max_col,
            'print_area': print_area,
            'page_margins': page_margins,
            'used_range': f"1:{max_row}, 1:{max_col}"
        })
        
    except Exception as e:
        current_app.logger.error(f"Preview error: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Preview failed: {str(e)}'
        }), 500

@templates_bp.route('/api/<template_name>/preview-fill', methods=['POST'])
def preview_fill_template(template_name):
    """Preview template with filled data before generating final file"""
    try:
        # Get form data
        data = request.get_json() or {}
        
        # Check if this is a Nike Form sheet
        if template_name in ['CPTPP CANADA', 'CPTPP MEXICO', 'D', 'AANZ', 'AK', 'CONG VAN ASI', 'CONG VAN CTH']:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], 'MergeDataNike.xlsm')
            sheet_name = template_name
        else:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], template_name)
            sheet_name = None
        
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'message': 'Template not found'}), 404
        
        from openpyxl import load_workbook
        import re
        
        wb = load_workbook(template_path, data_only=False)
        
        # Select the appropriate worksheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Preview data with filled values
        preview_data = []
        max_row = min(ws.max_row, 20)
        max_col = min(ws.max_column, 15)
        
        placeholder_pattern = re.compile(r'\{([^}]+)\}')
        
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell_value = cell.value if cell.value is not None else ""
                original_value = str(cell_value)
                
                # Replace placeholders with actual data
                if isinstance(cell_value, str) and '{' in cell_value and '}' in cell_value:
                    def replace_placeholder(match):
                        placeholder = match.group(1)
                        return str(data.get(placeholder, match.group(0)))
                    
                    filled_value = placeholder_pattern.sub(replace_placeholder, cell_value)
                    is_changed = filled_value != original_value
                else:
                    filled_value = original_value
                    is_changed = False
                
                row_data.append({
                    'original': original_value,
                    'filled': filled_value,
                    'is_changed': is_changed,
                    'address': cell.coordinate
                })
            preview_data.append(row_data)
        
        wb.close()
        
        return jsonify({
            'success': True,
            'template_name': template_name,
            'sheet_name': sheet_name or ws.title,
            'preview_data': preview_data,
            'total_rows': ws.max_row,
            'total_cols': ws.max_column
        })
        
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Preview failed: {str(e)}'
        }), 500

@templates_bp.route('/generate')
def generate_form():
    """Form to generate templates with data"""
    # Get invoice number from query params
    invoice_number = request.args.get('invoice_number', '')
    
    # Get available templates
    template_dir = current_app.config['TEMPLATE_FOLDER']
    templates = []
    for filename in os.listdir(template_dir):
        if filename.endswith(('.xlsx', '.xlsm', '.xls')):
            templates.append(filename)
    
    # Get invoice data if provided
    invoice_data = {}
    if invoice_number:
        invoice = get_invoice_by_number(invoice_number)
        if invoice and invoice.get('data_json'):
            try:
                invoice_data = json.loads(invoice['data_json'])
            except:
                pass
    
    return render_template('templates/generate.html', 
                         templates=templates, 
                         invoice_number=invoice_number,
                         invoice_data=invoice_data)

@templates_bp.route('/api/generate', methods=['POST'])
def api_generate_template():
    """API endpoint to generate filled templates"""
    try:
        data = request.get_json()
        template_name = data.get('template_name')
        invoice_number = data.get('invoice_number')
        custom_data = data.get('data', {})
        
        if not template_name:
            return jsonify({'success': False, 'message': 'Template name is required'}), 400
        
        # Check if this is a Nike Form sheet
        if template_name in ['CPTPP CANADA', 'CPTPP MEXICO', 'D', 'AANZ', 'AK', 'CONG VAN ASI', 'CONG VAN CTH']:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], 'MergeDataNike.xlsm')
            sheet_name = template_name
            base_name = template_name.replace(' ', '_')
        else:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], template_name)
            sheet_name = None
            base_name = os.path.splitext(template_name)[0]
            
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'message': 'Template not found'}), 404
        
        # Get data from various sources
        final_data = {}
        
        # 1. From database if invoice number provided
        if invoice_number:
            invoice = get_invoice_by_number(invoice_number)
            if invoice and invoice.get('data_json'):
                try:
                    final_data.update(json.loads(invoice['data_json']))
                except:
                    pass
        
        # 2. Override with custom data
        final_data.update(custom_data)
        
        # Generate output filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"{base_name}_{invoice_number or 'filled'}_{timestamp}.xlsx"
        output_path = os.path.join(current_app.config['OUTPUT_FOLDER'], output_filename)
        
        # Create output directory
        os.makedirs(current_app.config['OUTPUT_FOLDER'], exist_ok=True)
        
        # Generate filled template with preserved formatting
        template_manager = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER'])
        success, result = template_manager.create_filled_template(template_path, final_data, output_path, sheet_name)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Template generated successfully',
                'output_filename': output_filename,
                'replacements_made': result,
                'download_url': f'/templates/download/{output_filename}'
            })
        else:
            return jsonify({'success': False, 'message': f'Generation failed: {result}'}), 500
            
    except Exception as e:
        current_app.logger.error(f"Template generation error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@templates_bp.route('/download/<filename>')
def download_template(filename):
    """Download generated template"""
    try:
        file_path = os.path.join(current_app.config['OUTPUT_FOLDER'], filename)
        if not os.path.exists(file_path):
            flash('File not found', 'error')
            return redirect(url_for('templates.list_templates'))
        
        return send_file(file_path, as_attachment=True, download_name=filename)
    except Exception as e:
        current_app.logger.error(f"Download error: {str(e)}")
        flash('Download failed', 'error')
        return redirect(url_for('templates.list_templates'))

@templates_bp.route('/api/upload', methods=['POST'])
def api_upload_template():
    """Upload new template file"""
    try:
        if 'template_file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['template_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file type. Only Excel files allowed.'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], filename)
        
        # Check if file already exists
        if os.path.exists(template_path):
            return jsonify({'success': False, 'message': 'Template with this name already exists'}), 400
        
        file.save(template_path)
        
        # Validate template
        template_manager = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER'])
        errors, warnings = template_manager.validate_template(template_path)
        
        if errors:
            # Remove invalid file
            os.remove(template_path)
            return jsonify({
                'success': False, 
                'message': 'Invalid template file', 
                'errors': errors
            }), 400
        
        # Get template info
        info = template_manager.get_template_info(template_path)
        
        return jsonify({
            'success': True,
            'message': 'Template uploaded successfully',
            'template': info,
            'warnings': warnings
        })
        
    except Exception as e:
        current_app.logger.error(f"Template upload error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@templates_bp.route('/api/<template_name>/placeholders')
def api_get_placeholders(template_name):
    """Get placeholders from a template"""
    try:
        # Check if this is a Nike Form sheet
        if template_name in ['CPTPP CANADA', 'CPTPP MEXICO', 'D', 'AANZ', 'AK', 'CONG VAN ASI', 'CONG VAN CTH']:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], 'MergeDataNike.xlsm')
            sheet_name = template_name
        else:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], template_name)
            sheet_name = None
            
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'message': 'Template not found'}), 404
        
        template_manager = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER'])
        placeholders = template_manager.scan_template_placeholders(template_path, sheet_name)
        
        return jsonify({
            'success': True,
            'placeholders': placeholders,
            'count': len(placeholders),
            'sheet_name': sheet_name
        })
        
    except Exception as e:
        current_app.logger.error(f"Get placeholders error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@templates_bp.route('/api/batch-generate', methods=['POST'])
def api_batch_generate():
    """Generate multiple templates for multiple invoices"""
    try:
        data = request.get_json()
        template_names = data.get('template_names', [])
        invoice_numbers = data.get('invoice_numbers', [])
        
        if not template_names:
            return jsonify({'success': False, 'message': 'No templates specified'}), 400
        
        if not invoice_numbers:
            return jsonify({'success': False, 'message': 'No invoices specified'}), 400
        
        template_manager = NikeTemplateManager(current_app.config['TEMPLATE_FOLDER'])
        results = []
        
        for template_name in template_names:
            template_path = os.path.join(current_app.config['TEMPLATE_FOLDER'], template_name)
            if not os.path.exists(template_path):
                results.append({
                    'template': template_name,
                    'status': 'error',
                    'message': 'Template not found'
                })
                continue
            
            for invoice_number in invoice_numbers:
                # Get invoice data
                invoice = get_invoice_by_number(invoice_number)
                if not invoice:
                    results.append({
                        'template': template_name,
                        'invoice': invoice_number,
                        'status': 'error',
                        'message': 'Invoice not found'
                    })
                    continue
                
                # Get data
                invoice_data = {}
                if invoice.get('data_json'):
                    try:
                        invoice_data = json.loads(invoice['data_json'])
                    except:
                        pass
                
                # Generate filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name = os.path.splitext(template_name)[0]
                output_filename = f"{base_name}_{invoice_number}_{timestamp}.xlsx"
                output_path = os.path.join(current_app.config['OUTPUT_FOLDER'], output_filename)
                
                # Generate template
                success, result = template_manager.create_filled_template(
                    template_path, invoice_data, output_path
                )
                
                if success:
                    results.append({
                        'template': template_name,
                        'invoice': invoice_number,
                        'status': 'success',
                        'output_filename': output_filename,
                        'replacements_made': result
                    })
                else:
                    results.append({
                        'template': template_name,
                        'invoice': invoice_number,
                        'status': 'error',
                        'message': str(result)
                    })
        
        # Summary
        success_count = len([r for r in results if r['status'] == 'success'])
        total_count = len(results)
        
        return jsonify({
            'success': True,
            'message': f'Batch generation completed: {success_count}/{total_count} successful',
            'results': results,
            'summary': {
                'total': total_count,
                'successful': success_count,
                'failed': total_count - success_count
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Batch generation error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@templates_bp.route('/test-preview')
def test_preview():
    """Test preview page"""
    return render_template('test_preview.html')

@templates_bp.route('/simple-test')
def simple_test():
    """Simple jQuery test page"""
    return render_template('simple_test.html')