# -*- coding: utf-8 -*-
"""
Template Management System
Handles Excel templates with placeholder replacement similar to VBA logic
"""

import os
import re
import json
import logging
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy

logger = logging.getLogger(__name__)

class TemplateManager:
    """Manages Excel templates and placeholder replacement"""
    
    def __init__(self, template_folder):
        self.template_folder = template_folder
        self.placeholders_pattern = re.compile(r'\[([^\]]+)\]')
        
    def scan_template_placeholders(self, template_path):
        """Scan template file for placeholders"""
        placeholders = set()
        
        try:
            wb = load_workbook(template_path, data_only=False)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Scan cells for placeholders
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            found = self.placeholders_pattern.findall(cell.value)
                            placeholders.update(found)
                
                # Scan shapes/textboxes for placeholders (skip if not available)
                try:
                    if hasattr(ws, '_shapes') and ws._shapes:
                        for shape in ws._shapes:
                            if hasattr(shape, 'text') and shape.text:
                                found = self.placeholders_pattern.findall(shape.text)
                                placeholders.update(found)
                except:
                    pass  # Skip shapes scanning if not supported
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Error scanning template {template_path}: {str(e)}")
            
        return sorted(list(placeholders))
    
    def replace_placeholders_in_sheet(self, worksheet, data_dict):
        """Replace placeholders in a worksheet with actual data"""
        replacements_made = 0
        
        try:
            # Replace in cells
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        new_value = self._replace_placeholders_in_text(original_value, data_dict)
                        if new_value != original_value:
                            cell.value = new_value
                            replacements_made += 1
            
            # Replace in shapes/drawings (optional if available)
            try:
                if hasattr(worksheet, '_shapes') and worksheet._shapes:
                    for shape in worksheet._shapes:
                        if hasattr(shape, 'text') and shape.text:
                            original_text = shape.text
                            new_text = self._replace_placeholders_in_text(original_text, data_dict)
                            if new_text != original_text:
                                shape.text = new_text
                                replacements_made += 1
            except Exception:
                pass
                        
        except Exception as e:
            logger.error(f"Error replacing placeholders in sheet: {str(e)}")
            
        return replacements_made
    
    def _replace_placeholders_in_text(self, text, data_dict):
        """Replace placeholders in text with values from data dictionary"""
        def replace_func(match):
            placeholder = match.group(1)
            return str(data_dict.get(placeholder, match.group(0)))
        
        return self.placeholders_pattern.sub(replace_func, text)
    
    def create_filled_template(self, template_path, data_dict, output_path):
        """Create a new file with placeholders filled from template"""
        try:
            # Load template
            wb = load_workbook(template_path, keep_vba=False)
            total_replacements = 0
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                replacements = self.replace_placeholders_in_sheet(ws, data_dict)
                total_replacements += replacements
                
                # Apply page setup similar to VBA
                self._apply_page_setup(ws)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save filled template
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Template filled successfully: {total_replacements} replacements made")
            return True, total_replacements
            
        except Exception as e:
            error_msg = f"Error creating filled template: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def _apply_page_setup(self, worksheet):
        """Apply page setup settings similar to VBA EnsurePageSetup"""
        try:
            # Set page setup for printing
            worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
            worksheet.page_setup.fitToPage = True
            worksheet.page_setup.fitToHeight = 1
            worksheet.page_setup.fitToWidth = 1
            
            # Set margins (in inches)
            worksheet.page_margins.left = 0.5
            worksheet.page_margins.right = 0.5
            worksheet.page_margins.top = 0.5
            worksheet.page_margins.bottom = 0.5
            worksheet.page_margins.header = 0.3
            worksheet.page_margins.footer = 0.3
            
        except Exception as e:
            logger.error(f"Error applying page setup: {str(e)}")
    
    def export_to_pdf(self, excel_path, pdf_path):
        """Export Excel file to PDF (would need additional libraries)"""
        # Note: This would require additional PDF conversion libraries
        # For now, we'll return the excel path and let the frontend handle it
        logger.info(f"PDF export requested: {excel_path} -> {pdf_path}")
        return excel_path
    
    def get_template_info(self, template_path):
        """Get information about a template file"""
        info = {
            'path': template_path,
            'name': os.path.basename(template_path),
            'exists': os.path.exists(template_path),
            'size': 0,
            'modified': None,
            'created_at': None,
            'display_name': None,
            'file_size_mb': 0.0,
            'usage_count': 0,
            'sheets': [],
            'placeholders': []
        }
        
        if info['exists']:
            try:
                # File stats
                stat = os.stat(template_path)
                info['size'] = stat.st_size
                info['modified'] = datetime.fromtimestamp(stat.st_mtime)
                info['created_at'] = info['modified']
                info['display_name'] = os.path.splitext(info['name'])[0]
                info['file_size_mb'] = round(info['size'] / (1024 * 1024), 2)
                
                # Workbook info with per-sheet details
                wb = load_workbook(template_path, data_only=False)
                all_placeholders = set()
                sheets_info = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Gather placeholders with coordinates for this sheet
                    sheet_placeholders = []
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                for ph in self.placeholders_pattern.findall(cell.value):
                                    sheet_placeholders.append({
                                        'cell': cell.coordinate,
                                        'placeholder': ph
                                    })
                                    all_placeholders.add(ph)
                    # Build sheet info
                    sheets_info.append({
                        'name': sheet_name,
                        'row_count': ws.max_row,
                        'col_count': ws.max_column,
                        'placeholders': sheet_placeholders
                    })
                info['sheets'] = sheets_info
                info['placeholders'] = sorted(list(all_placeholders))
                wb.close()
                
            except Exception as e:
                logger.error(f"Error getting template info for {template_path}: {str(e)}")
                
        return info
    
    def validate_template(self, template_path):
        """Validate template file"""
        errors = []
        warnings = []
        
        if not os.path.exists(template_path):
            errors.append("Template file does not exist")
            return errors, warnings
        
        try:
            wb = load_workbook(template_path, data_only=False)
            
            # Check if template has any sheets
            if not wb.sheetnames:
                errors.append("Template has no worksheets")
            
            # Check for placeholders
            placeholders = self.scan_template_placeholders(template_path)
            if not placeholders:
                warnings.append("No placeholders found in template")
            
            # Check for common required placeholders
            required_fields = ['Invoice Number', 'Date', 'Total Cartons', 'Marks']
            missing_required = [field for field in required_fields if field not in placeholders]
            if missing_required:
                warnings.append(f"Missing common placeholders: {', '.join(missing_required)}")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Error reading template: {str(e)}")
        
        return errors, warnings

def create_sample_template(output_path, template_name="Sample CO Template"):
    """Create a sample Certificate of Origin template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate of Origin"
    
    # Sample template structure
    template_data = [
        ["", "", "", "CERTIFICATE OF ORIGIN", "", "", ""],
        ["", "", "", "Form A", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Invoice Number:", "{Invoice Number}", "", "Date:", "{Date}", "", ""],
        ["", "", "", "", "", "", ""],
        ["Exporter:", "", "", "Consignee:", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Description of Goods:", "", "", "", "", "", ""],
        ["{DESCRIPTION}", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Marks and Numbers:", "", "Origin Criterion:", "", "", "", ""],
        ["{Marks}", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Total Cartons:", "{Total Cartons}", "", "In Words:", "{Total Cartons In Words}", "", ""],
        ["", "", "", "", "", "", ""],
        ["Destination:", "{DEST}", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Plant:", "{Plant}", "", "Material:", "{Material}", "", ""],
        ["PO#:", "{PO}", "", "Reference PO#:", "{Reference PO#}", "", ""],
    ]
    
    # Fill template
    for row_idx, row_data in enumerate(template_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Style headers
            if row_idx <= 2 and cell_value:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center')
            elif ":" in str(cell_value) and not cell_value.startswith("{"):
                cell.font = Font(bold=True)
    
    # Merge title cells
    ws.merge_cells('D1:F1')
    ws.merge_cells('D2:F2')
    
    # Save template
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    
    return output_path

if __name__ == '__main__':
    # Create sample template for testing
    template_path = "templates/sample_co_template.xlsx"
    create_sample_template(template_path)
    print(f"Sample template created: {template_path}")
    
    # Test template manager
    tm = TemplateManager("templates")
    info = tm.get_template_info(template_path)
    print(f"Template info: {json.dumps(info, default=str, indent=2)}")
    
    # Test placeholder replacement
    test_data = {
        'Invoice Number': 'A1234567',
        'Date': '10/10/2025',
        'DESCRIPTION': 'SKULL CAP',
        'Marks': 'NIKE 1234567 VIETNAM',
        'Total Cartons': '100',
        'Total Cartons In Words': 'ONE HUNDRED',
        'DEST': 'United States',
        'Plant': 'VN01',
        'Material': 'MAT123',
        'PO': '1234567',
        'Reference PO#': '7654321'
    }
    
    success, result = tm.create_filled_template(template_path, test_data, "outputs/filled_sample.xlsx")
    print(f"Template filling: {'Success' if success else 'Failed'} - {result}")