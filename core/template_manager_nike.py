# -*- coding: utf-8 -*-
"""
Template Management System for Nike Export
Based on original ExportNike_XuatKhau_AllInOne.py logic
"""

import os
import re
import json
import logging
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy

logger = logging.getLogger(__name__)

class NikeTemplateManager:
    """Manages Excel templates with Nike-specific field mapping"""
    
    def __init__(self, template_folder):
        self.template_folder = template_folder
        self.placeholders_pattern = re.compile(r'\{([^}]+)\}')
        
        # Field mapping from original script
        self.field_mapping = {
            "inv": "Invoice Number",
            "date": "Date", 
            "po": "PO",
            "refpo": "Reference PO#",
            "itemseq": "Item Seq.",
            "material": "Material",
            "desc": "Desc",
            "custship": "Customer Ship To #",
            "plant": "Plant",
            "gross": "Total Gross Kgs",
            "cartons": "Total Cartons",
            "units": "Total Units",
            "marks": "Marks",
            "description": "DESCRIPTION",
            "dest": "DEST"
        }
    
    def header_key(self, s):
        """Convert header text to normalized key"""
        return re.sub(r"[^a-z0-9#]+", " ", str(s).lower()).strip()
    
    def build_header_map_and_get_cols(self, ws, header_row: int, verbose: bool = False):
        """
        Build header mapping and return column indices
        Based on original logic with fallback to default positions
        """
        header = {}
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(row=header_row, column=c).value
            if raw is not None and str(raw).strip() != "":
                header[self.header_key(raw)] = c

        def pick(*aliases):
            for a in aliases:
                if a in header:
                    return header[a]
            return None

        cols = {
            "inv":         pick("invoice number", "invoice no", "invoice"),
            "date":        pick("date"),
            "po":          pick("po", "po#", "purchase order", "purchase order #"),
            "refpo":       pick("reference po#", "reference po #", "reference po", "po# reference"),
            "itemseq":     pick("item seq", "item seq.", "item", "seq"),
            "material":    pick("material"),
            "desc":        pick("desc", "description short", "product desc"),
            "custship":    pick("customer ship to #", "customer ship to", "ship to #"),
            "plant":       pick("plant"),
            "gross":       pick("total gross kgs", "gross kgs"),
            "cartons":     pick("total cartons", "cartons"),
            "units":       pick("total units", "units"),
            "marks":       pick("marks"),
            "description": pick("description"),
            "dest":        pick("dest", "destination"),
        }

        def ensure(name, default_idx):
            v = cols.get(name)
            if not isinstance(v, int) or v < 1:
                cols[name] = default_idx

        # Set default column positions if not found
        ensure("inv", 1);  ensure("date", 2);  ensure("po", 3);  ensure("refpo", 4)
        ensure("itemseq", 5);  ensure("material", 6);  ensure("desc", 7);  ensure("custship", 8)
        ensure("plant", 9);  ensure("gross", 10);  ensure("cartons", 11);  ensure("units", 12)
        ensure("marks", 13); ensure("description", 14); ensure("dest", 15)

        if verbose:
            logger.info(f"Header mapping: {cols}")
        return cols
    
    def find_or_append_row(self, ws, inv_col: int, start_row: int, invoice: str) -> int:
        """Find existing row or append new one for invoice"""
        inv_key = re.sub(r"\s+", " ", str(invoice)).strip().upper()
        first_empty = None
        
        for r in range(start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=inv_col).value
            if v is None or str(v).strip() == "":
                first_empty = r
                break
            if re.sub(r"\s+", " ", str(v)).strip().upper() == inv_key:
                return r
        
        target = first_empty if first_empty else (ws.max_row + 1 if ws.max_row >= start_row else start_row)
        ws.cell(row=target, column=inv_col).value = invoice
        return target
    
    def fill_excel_row(self, ws, cols, start_row, inv, pl_row_dict, asi_desc, dest_override=None):
        """Fill Excel row with data from processing results"""
        r = self.find_or_append_row(ws, cols["inv"], start_row, inv)

        def setc(colkey, val):
            if val not in (None, ""):
                ws.cell(row=r, column=cols[colkey]).value = val

        # Fill basic data from PL
        setc("date",        pl_row_dict.get("Date"))
        setc("po",          pl_row_dict.get("PO"))
        setc("refpo",       pl_row_dict.get("Reference PO#"))
        setc("itemseq",     pl_row_dict.get("Item Seq."))
        setc("material",    pl_row_dict.get("Material"))
        setc("desc",        pl_row_dict.get("Desc"))
        setc("custship",    pl_row_dict.get("Customer Ship To #"))
        setc("plant",       pl_row_dict.get("Plant"))
        setc("gross",       pl_row_dict.get("Total Gross Kgs"))
        setc("cartons",     pl_row_dict.get("Total Cartons"))
        setc("units",       pl_row_dict.get("Total Units"))
        setc("marks",       pl_row_dict.get("Marks"))

        # Set destination
        if dest_override:
            setc("dest", dest_override)
        elif pl_row_dict.get("DEST"):
            setc("dest", pl_row_dict.get("DEST"))
        
        # Set description from ASI/BOOKING
        if asi_desc:
            setc("description", asi_desc)
        
        return r
    
    def scan_template_placeholders(self, template_path, sheet_name=None):
        """Scan template file for placeholders"""
        placeholders = set()
        
        try:
            wb = load_workbook(template_path, data_only=False)
            
            # If specific sheet name is provided, scan only that sheet
            if sheet_name and sheet_name in wb.sheetnames:
                sheets_to_scan = [sheet_name]
            else:
                sheets_to_scan = wb.sheetnames
            
            for sheet in sheets_to_scan:
                ws = wb[sheet]
                
                # Scan cells for placeholders
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            found = self.placeholders_pattern.findall(cell.value)
                            placeholders.update(found)
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Error scanning template {template_path}, sheet {sheet_name}: {str(e)}")
            
        return sorted(list(placeholders))
    
    def fill_template(self, template_path, sheet_name, data_dict, output_folder="outputs"):
        """
        Fill template with data and return output file path
        
        Args:
            template_path: Path to template file
            sheet_name: Name of sheet to fill
            data_dict: Dictionary with data to fill
            output_folder: Folder to save output file
            
        Returns:
            str: Path to filled template file
        """
        try:
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            template_name = os.path.splitext(os.path.basename(template_path))[0]
            output_filename = f"{template_name}_{sheet_name}_{timestamp}.xlsx"
            output_path = os.path.join(output_folder, output_filename)
            
            # Ensure output directory exists
            os.makedirs(output_folder, exist_ok=True)
            
            # Load workbook
            wb = load_workbook(template_path, keep_vba=True)
            
            # Check if sheet exists
            if sheet_name not in wb.sheetnames:
                available_sheets = ", ".join(wb.sheetnames)
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
            
            ws = wb[sheet_name]
            
            # Replace placeholders
            replacements = self.replace_placeholders_in_sheet(ws, data_dict)
            
            # Save filled template
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Template filled successfully: {output_path} ({replacements} replacements)")
            return output_path
            
        except Exception as e:
            error_msg = f"Error filling template: {str(e)}"
            logger.error(error_msg)
            raise Exception(error_msg)
    
    def replace_placeholders_in_sheet(self, worksheet, data_dict):
        """Replace placeholders in a worksheet with actual data"""
        replacements_made = 0
        
        try:
            # Replace in cells
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        new_value = self._replace_placeholders_in_text(original_value, data_dict)
                        if new_value != original_value:
                            cell.value = new_value
                            replacements_made += 1
                        
        except Exception as e:
            logger.error(f"Error replacing placeholders in sheet: {str(e)}")
            
        return replacements_made
    
    def _replace_placeholders_in_text(self, text, data_dict):
        """Replace placeholders in text with values from data dictionary"""
        def replace_func(match):
            placeholder = match.group(1)
            return str(data_dict.get(placeholder, match.group(0)))
        
        return self.placeholders_pattern.sub(replace_func, text)
    
    def create_filled_template(self, template_path, data_dict, output_path, sheet_name=None):
        """Create a new file with placeholders filled from template - preserves ALL formatting"""
        try:
            # Load template with keep_vba=True to preserve all formatting and macros
            wb = load_workbook(template_path, keep_vba=True, data_only=False)
            total_replacements = 0
            
            # If specific sheet name provided, process only that sheet
            if sheet_name and sheet_name in wb.sheetnames:
                sheets_to_process = [sheet_name]
            else:
                # Skip 'Data' sheet as requested - only process forms
                sheets_to_process = [s for s in wb.sheetnames if s != 'Data']
            
            # Process selected sheets only
            for sheet in sheets_to_process:
                ws = wb[sheet]
                replacements = self.replace_placeholders_in_sheet_preserve_format(ws, data_dict)
                total_replacements += replacements
                logger.info(f"Sheet '{sheet}': {replacements} replacements")
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save with ALL original formatting preserved
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Created filled template: {output_path} ({total_replacements} total replacements)")
            return True, f"Successfully created template with {total_replacements} replacements"
            
        except Exception as e:
            error_msg = f"Error creating filled template: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def replace_placeholders_in_sheet_preserve_format(self, worksheet, data_dict):
        """Replace placeholders in worksheet while preserving ALL original formatting"""
        replacements_made = 0
        
        try:
            # Iterate through all cells and replace only text content
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '{' in cell.value:
                        # Store original formatting
                        original_font = cell.font
                        original_fill = cell.fill
                        original_border = cell.border
                        original_alignment = cell.alignment
                        original_number_format = cell.number_format
                        
                        # Replace placeholders in text only
                        old_value = cell.value
                        new_value = self.replace_placeholders_in_text(old_value, data_dict)
                        
                        if old_value != new_value:
                            # Update cell value
                            cell.value = new_value
                            
                            # Restore ALL original formatting
                            cell.font = original_font
                            cell.fill = original_fill  
                            cell.border = original_border
                            cell.alignment = original_alignment
                            cell.number_format = original_number_format
                            
                            replacements_made += 1
                            logger.debug(f"Replaced in {cell.coordinate}: '{old_value}' -> '{new_value}'")
            
        except Exception as e:
            logger.error(f"Error replacing placeholders in sheet: {str(e)}")
            
        return replacements_made
    
    def process_nike_data(self, template_path, invoice_data, output_path, sheet_name="Data", start_row=8):
        """
        Process Nike data and fill template
        
        Args:
            template_path: Path to Excel template
            invoice_data: Dictionary with invoice information
            output_path: Path for output file
            sheet_name: Name of worksheet to update
            start_row: Starting row for data
        """
        try:
            # Load workbook
            wb = load_workbook(template_path, keep_vba=True)
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in template")
            
            ws = wb[sheet_name]
            header_row = start_row - 1
            cols = self.build_header_map_and_get_cols(ws, header_row, verbose=True)
            
            # Extract data components
            invoice_number = invoice_data.get("invoice_number", "")
            pl_data = invoice_data.get("pl_data", {})
            booking_data = invoice_data.get("booking_data", {})
            
            # Get ASI description
            asi_desc = booking_data.get("DESCRIPTION", "")
            
            # Fill row
            row_num = self.fill_excel_row(
                ws, cols, start_row, 
                invoice_number, pl_data, asi_desc
            )
            
            # Save workbook
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Successfully processed Nike data for invoice {invoice_number}")
            return True, f"Data filled in row {row_num}"
            
        except Exception as e:
            error_msg = f"Error processing Nike data: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def get_template_info(self, template_path):
        """Get information about a template file"""
        info = {
            'path': template_path,
            'name': os.path.basename(template_path),
            'exists': os.path.exists(template_path),
            'size': 0,
            'modified': None,
            'sheets': [],
            'placeholders': []
        }
        
        if info['exists']:
            try:
                # File stats
                stat = os.stat(template_path)
                info['size'] = stat.st_size
                info['modified'] = datetime.fromtimestamp(stat.st_mtime)
                
                # Workbook info
                wb = load_workbook(template_path, data_only=False)
                info['sheets'] = wb.sheetnames
                info['placeholders'] = self.scan_template_placeholders(template_path)
                wb.close()
                
            except Exception as e:
                logger.error(f"Error getting template info for {template_path}: {str(e)}")
                
        return info
    
    def validate_template(self, template_path):
        """Validate template file"""
        errors = []
        warnings = []
        
        try:
            if not os.path.exists(template_path):
                errors.append("Template file does not exist")
                return errors, warnings
            
            # Try to load workbook
            wb = load_workbook(template_path, data_only=False)
            
            # Check if has at least one sheet
            if not wb.sheetnames:
                errors.append("Template has no worksheets")
            
            # Check for placeholders
            placeholders = self.scan_template_placeholders(template_path)
            if not placeholders:
                warnings.append("No placeholders found in template")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Error validating template: {str(e)}")
        
        return errors, warnings

def create_sample_nike_template(output_path, template_name="Nike CO Template"):
    """Create a sample Nike Certificate of Origin template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate of Origin"
    
    # Nike-specific template structure
    template_data = [
        ["", "", "", "CERTIFICATE OF ORIGIN", "", "", ""],
        ["", "", "", "Form A", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Invoice Number:", "{Invoice Number}", "", "Date:", "{Date}", "", ""],
        ["", "", "", "", "", "", ""],
        ["Exporter:", "", "", "Consignee:", "", "", ""],
        ["NIKE VIETNAM LLC", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Description of Goods:", "", "", "", "", "", ""],
        ["{DESCRIPTION}", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Marks and Numbers:", "", "Origin Criterion:", "", "", "", ""],
        ["{Marks}", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Total Cartons:", "{Total Cartons}", "", "In Words:", "{Total Cartons In Words}", "", ""],
        ["", "", "", "", "", "", ""],
        ["Destination:", "{DEST}", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["Plant:", "{Plant}", "", "Material:", "{Material}", "", ""],
        ["PO#:", "{PO}", "", "Reference PO#:", "{Reference PO#}", "", ""],
    ]
    
    # Fill template
    for row_idx, row_data in enumerate(template_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Style headers
            if row_idx <= 2 and cell_value:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center')
            elif ":" in str(cell_value) and not cell_value.startswith("{"):
                cell.font = Font(bold=True)
    
    # Merge title cells
    ws.merge_cells('D1:F1')
    ws.merge_cells('D2:F2')
    
    # Save template
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    
    return output_path

if __name__ == '__main__':
    # Test the template manager
    template_path = "templates/nike_co_template.xlsx"
    create_sample_nike_template(template_path)
    print(f"Nike template created: {template_path}")
    
    # Test template manager
    tm = NikeTemplateManager("templates")
    info = tm.get_template_info(template_path)
    print(f"Template info: {json.dumps(info, default=str, indent=2)}")